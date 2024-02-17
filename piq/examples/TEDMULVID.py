import argparse
import torch
import piq
from openpyxl import Workbook, load_workbook
from skimage.io import imread
import os

# Function to calculate PSNR, SSIM, FSIM, VSI, DISTS, BRISQUE, and CLIP-IQA metrics
@torch.no_grad()
def main(img, gt):
    metrics = []
    # Load RGB images and corresponding ground truth images
    x = torch.tensor(imread(img)).permute(2, 0, 1)[None, ...] / 255.
    y = torch.tensor(imread(gt)).permute(2, 0, 1)[None, ...] / 255.

    # Use GPU for faster computations if available
    if torch.cuda.is_available():
        x = x.cuda()
        y = y.cuda()

    # Calculate and append different image quality metrics
    psnr_index = piq.psnr(x, y, data_range=1., reduction='none')
    metrics.append(psnr_index.item())

    ssim_index = piq.ssim(x, y, data_range=1.)
    metrics.append(ssim_index.item())

    fsim_index: torch.Tensor = piq.fsim(x, y, data_range=1., reduction='none')
    metrics.append(fsim_index.item())

    vsi_index: torch.Tensor = piq.vsi(x, y, data_range=1.)
    metrics.append(vsi_index.item())

    dists_loss = piq.DISTS(reduction='none')(x, y)
    metrics.append(dists_loss.item())

    brisque_index: torch.Tensor = piq.brisque(x, data_range=1., reduction='none')
    metrics.append(brisque_index.item())

    clip_iqa_index: torch.Tensor = piq.CLIPIQA(data_range=1.).to(x.device)(x)
    metrics.append(clip_iqa_index.item())

    return metrics

def numeric_sort_key(s):
    """
    Key function for sorting strings containing numeric and non-numeric values.
    Converts the string to an integer if it's numeric, otherwise leaves it as is.
    """
    return f"{int(s):010d}" if s.isdigit() else s

def read_png(scene_path, target_folder, gt_path, excel_path):
    for root, dirs, files in os.walk(scene_path):
        # Sort directories using the numeric sort key
        dirs.sort(key=numeric_sort_key)
        if target_folder in dirs:
            full_path = os.path.join(root, target_folder)
            video_num = os.path.basename(root)
            # Process each file in the sorted list
            PSNR, SSIM, FSIM, VSI, DISTS, BRISQUE, CLIP = ([] for i in range(7)) # metric values for each video in current scene
            for file in sorted(os.listdir(full_path), key=lambda x: int(x.split('.')[0]) if x.split('.')[0].isdigit() else x):
                img_len = len(os.listdir(full_path))
                if file.endswith(".png"):
                    image_path = os.path.join(full_path, file)
                    if image_path:
                        metrics = main(image_path, gt_path) # calculate metrics
                        PSNR.append(metrics[0])
                        SSIM.append(metrics[1])
                        FSIM.append(metrics[2])
                        VSI.append(metrics[3])
                        DISTS.append(metrics[4])
                        BRISQUE.append(metrics[5])
                        CLIP.append(metrics[6])
                        # Calculate average metrics when the last image is reached
                        if os.path.basename(image_path) == str(img_len) + ".png":
                            avarage_values = [sum(metric) / len(metric) for metric in [PSNR, SSIM, FSIM, VSI, DISTS, BRISQUE, CLIP]]
                            append_metrics_to_excel(avarage_values, excel_path, target_folder, video_num)
                    else:
                        print(f"Unable to read or empty file: {image_path}")

def create_or_load_excel(file_path):
    # If the file already exists, load the file
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
    else:
        # If the file does not exist, create a new workbook and add a worksheet
        workbook = Workbook()
        workbook.create_sheet(index=0)

    return workbook

def append_metrics_to_excel(metrics, file_path, target_folder, video_num):
    # Load or create the workbook
    wb = create_or_load_excel(file_path)
    ws = wb.active

    # Determine the new row for adding new data
    new_row = ws.max_row + 1

    # Write the target folder name and video number in the first two columns
    ws.cell(row=new_row, column=1, value=target_folder)
    ws.cell(row=new_row, column=2, value=video_num)

    # Starting from the third column, write the rounded metrics
    for col, value in enumerate(metrics, start=3):
        ws.cell(row=new_row, column=col, value=round(value, 4))  # Round to 4 decimal places

    # Save the Excel workbook
    wb.save(file_path)

if __name__ == '__main__':

    # Define the command-line arguments
    parser = argparse.ArgumentParser(description='Low-Light Image Enhancement and Evaluation')
    parser.add_argument('--method', type=str, required=True, help='Enhancement method')
    parser.add_argument('--scene', type=str, required=True, help='Scene folder name')
    parser.add_argument('--gt', type=str, required=True, help='Ground truth image file name')
    parser.add_argument('--excel', type=str, required=True, help='Excel file path')

    # Parse the arguments
    args = parser.parse_args()

    # Base directory path
    base_dir = os.getcwd()

    # Construct file paths
    scene_path = os.path.join(base_dir, "examples", args.scene)
    target_folder = args.method
    gt_path = os.path.join(base_dir, "examples", args.scene, "GT", args.gt)
    excel_path = os.path.join(base_dir, "examples", args.scene, args.excel)

    # Call the function to process images
    read_png(scene_path, target_folder, gt_path, excel_path)
